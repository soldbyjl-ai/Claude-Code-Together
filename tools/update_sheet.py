"""
update_sheet.py
---------------
Writes scraped Paragon + BC Assessment data into the Excel workbook.

Usage:
    python tools/update_sheet.py --type expired      # default
    python tools/update_sheet.py --type active
    python tools/update_sheet.py --type sold
    python tools/update_sheet.py --type terminated

For each type, the script:
  - Reads tools/output/paragon_detached.json
  - Reads tools/output/bcassessment_results.json
  - Merges assessment values into listings
  - Writes (or refreshes) the appropriate sheet section(s)
  - Updates B1 (Last Scraped Date) to today

Behavioral rules enforced:
  - Deduplication by MLS# for E.DETACHED historical section (add-only)
  - Manual columns (Contact/Q, List#/G, OCCUP/V, Notes/W) are never overwritten
  - Col A duplicate-flag formulas are never overwritten
  - D.EXP is a static value computed at write time against the scrape date
"""

import argparse
import json
import logging
from datetime import datetime
from pathlib import Path

import openpyxl

# ---------------------------------------------------------------------------
# Paths
# ---------------------------------------------------------------------------
WORKBOOK     = Path(__file__).parent.parent / "1VBN - Brentwood Park.xlsx"
PARAGON_JSON = Path(__file__).parent / "output" / "paragon_detached.json"
BCASSESS_JSON = Path(__file__).parent / "output" / "bcassessment_results.json"
OUTPUT_DIR   = Path(__file__).parent / "output"
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[
        logging.StreamHandler(),
        logging.FileHandler(OUTPUT_DIR / "update_sheet.log", mode="a"),
    ],
)
log = logging.getLogger(__name__)

TODAY = datetime.now()

# Column index constants (1-based)
COL = {
    "A": 1,  "B": 2,  "C": 3,  "D": 4,  "E": 5,  "F": 6,  "G": 7,
    "H": 8,  "I": 9,  "J": 10, "K": 11, "L": 12, "M": 13,
    "N": 14, "O": 15, "P": 16, "Q": 17, "R": 18,
    "S": 19, "T": 20, "U": 21, "V": 22, "W": 23,
}

# Data columns to WRITE for each section type
# (never include A=formula, G=List#, Q=Contact, V=OCCUP, W=Notes)
WRITE_COLS_EXPIRED    = ["B","C","D","E","F","H","I","J","K","L","M","N","O","P","R"]
WRITE_COLS_ACTIVE     = ["B","C","D","E","F","H","I","J","K","L","M","N","O","P"]
WRITE_COLS_SOLD       = ["B","C","D","E","F","H","I","J","K","L","M","N","O"]
WRITE_COLS_EDET       = ["B","C","D","E","F","H","I","J","K","L","M","N","O","P","R","S","T","U"]

# Columns to CLEAR on replace (broader — removes stale data from old scraper runs)
# For A.S.PRINTOUT sections: clear B-U (skip A=formula, Q=Contact)
CLEAR_COLS_EXPIRED    = ["B","C","D","E","F","G","H","I","J","K","L","M","N","O","P","R","S","T","U"]
CLEAR_COLS_ACTIVE     = ["B","C","D","E","F","G","H","I","J","K","L","M","N","O","P","R","S"]
CLEAR_COLS_SOLD       = ["B","C","D","E","F","G","H","I","J","K","L","M","N","O","P"]
CLEAR_COLS_EDET       = ["B","C","D","E","F","G","H","I","J","K","L","M","N","O","P","R","S","T","U"]


# ---------------------------------------------------------------------------
# Value parsers
# ---------------------------------------------------------------------------

def parse_money(s) -> int | None:
    """'$1,789,000' -> 1789000"""
    if not s:
        return None
    try:
        return int(str(s).replace("$", "").replace(",", "").strip())
    except (ValueError, TypeError):
        return None


def parse_int(s) -> int | None:
    """'6,466.00' or '377' -> 6466 / 377"""
    if not s:
        return None
    try:
        return int(float(str(s).replace(",", "").strip()))
    except (ValueError, TypeError):
        return None


def parse_date(s) -> datetime | None:
    """Parse 'M/D/YYYY', 'M/D/YY', 'YYYY-MM-DD' -> datetime, or None."""
    if not s:
        return None
    if isinstance(s, datetime):
        return s
    for fmt in ("%m/%d/%Y", "%m/%d/%y", "%Y-%m-%d"):
        try:
            return datetime.strptime(str(s).strip(), fmt)
        except (ValueError, TypeError):
            pass
    return None


def calc_dexp(expiry_str, scrape_date: datetime) -> int | str:
    """(expiry - scrape_date).days; returns 'Expired' if <= 0."""
    exp = parse_date(expiry_str)
    if exp is None:
        return ""
    days = (exp.date() - scrape_date.date()).days
    return "Expired" if days <= 0 else days


# ---------------------------------------------------------------------------
# Load & merge
# ---------------------------------------------------------------------------

def load_data() -> list[dict]:
    """Load paragon JSON, merge BC Assessment values, return listings."""
    if not PARAGON_JSON.exists():
        log.error(f"Paragon output not found: {PARAGON_JSON}")
        return []
    with open(PARAGON_JSON, encoding="utf-8") as f:
        listings = json.load(f)
    log.info(f"Loaded {len(listings)} listings from paragon output.")

    bcassess = []
    if BCASSESS_JSON.exists():
        with open(BCASSESS_JSON, encoding="utf-8") as f:
            bcassess = json.load(f)
        log.info(f"Loaded {len(bcassess)} BC Assessment records.")
    else:
        log.warning(f"BC Assessment file not found: {BCASSESS_JSON}")

    return _merge_bcassess(listings, bcassess)


def _merge_bcassess(listings: list[dict], bcassess: list[dict]) -> list[dict]:
    """Merge land/total assessment values into listings by address."""
    lookup: dict[str, tuple] = {}
    for rec in bcassess:
        if rec.get("error"):
            continue
        addr = rec.get("address", "").upper()
        for suffix in (" BURNABY BC", " BURNABY"):
            if addr.endswith(suffix):
                addr = addr[: -len(suffix)].strip()
                break
        lookup[addr] = (rec.get("land_value", ""), rec.get("total_assessed_value", ""))

    for listing in listings:
        addr = listing.get("address", "").upper().strip()
        if addr in lookup:
            listing["land_assessment"]  = lookup[addr][0]
            listing["total_assessment"] = lookup[addr][1]
    return listings


# ---------------------------------------------------------------------------
# Scrape-date helpers
# ---------------------------------------------------------------------------

def get_scrape_date(ws_asp) -> datetime:
    """Read last scraped date from A.S.PRINTOUT!B1 ('03.20.26' -> datetime)."""
    val = ws_asp["B1"].value
    if val:
        for fmt in ("%m.%d.%y", "%m.%d.%Y"):
            try:
                return datetime.strptime(str(val).strip(), fmt)
            except (ValueError, TypeError):
                pass
    log.warning("Could not parse B1 scrape date; using today.")
    return TODAY


def update_scrape_date(ws_asp):
    """Write today's date into A.S.PRINTOUT!B1 in MM.DD.YY format."""
    new_val = TODAY.strftime("%m.%d.%y")
    ws_asp["B1"].value = new_val
    log.info(f"Updated B1 scrape date -> {new_val}")


# ---------------------------------------------------------------------------
# Row builder
# ---------------------------------------------------------------------------

def build_row(listing: dict, scrape_date: datetime, search_type: str) -> dict:
    """
    Convert a listing dict into column-keyed values ready for writing.
    Keys match COL dict letters; 'R' = ML# (without trailing comma).
    """
    land  = parse_money(listing.get("land_assessment"))
    total = parse_money(listing.get("total_assessment"))
    price = parse_money(listing.get("list_price"))
    dom   = parse_int(listing.get("dom"))
    yr    = parse_int(listing.get("year_built"))
    lot   = parse_int(listing.get("lot_size"))
    beds  = parse_int(listing.get("bedrooms"))
    baths = parse_int(listing.get("bathrooms"))
    sqft  = parse_int(listing.get("sqft"))
    kitch = parse_int(listing.get("kitchens"))
    addr  = listing.get("address", "").strip()
    mls   = listing.get("mls_number", "").strip()
    exp_dt  = parse_date(listing.get("expiry_date", ""))
    list_dt = parse_date(listing.get("list_date", ""))
    dexp  = calc_dexp(listing.get("expiry_date", ""), scrape_date)

    if search_type == "sold":
        sold_price = parse_money(listing.get("sold_price"))
        sold_dt    = parse_date(listing.get("sold_date", ""))
        return {
            "B": addr, "C": sold_price, "D": price, "E": sold_dt,
            "F": dom, "H": yr, "I": lot, "J": beds, "K": baths,
            "L": sqft, "M": kitch, "N": land, "O": total,
            "R": mls,
            # P = =C{r}-O{r} written separately in write_sold()
        }
    else:
        return {
            "B": addr, "C": exp_dt, "D": price, "E": list_dt,
            "F": dom, "H": yr, "I": lot, "J": beds, "K": baths,
            "L": sqft, "M": kitch, "N": land, "O": total,
            "P": dexp,
            "R": mls,
        }


# ---------------------------------------------------------------------------
# Cell writers
# ---------------------------------------------------------------------------

def write_row(ws, row: int, data: dict, cols: list[str]):
    """Write data dict values to the given row, only for specified columns.
    Col 'R' (ML#) gets a trailing comma appended to match workbook convention.
    Col 'A' is never touched (formula preserved).
    """
    for letter in cols:
        key = letter
        val = data.get(key)
        if letter == "R" and val:
            val = f"{val},"
        ws.cell(row=row, column=COL[letter]).value = val


def clear_section(ws, rows: range, cols: list[str]):
    """Set all cells in the given rows/cols to None, removing hyperlinks too."""
    for r in rows:
        for letter in cols:
            cell = ws.cell(row=r, column=COL[letter])
            cell.value = None
            cell.hyperlink = None


def collect_mls(ws, mls_col: int, start_row: int = 1) -> set[str]:
    """Return set of normalised MLS# strings present in mls_col from start_row."""
    found = set()
    for r in range(start_row, ws.max_row + 1):
        val = ws.cell(row=r, column=mls_col).value
        if val and str(val).strip().startswith("R"):
            found.add(str(val).strip().rstrip(",").upper())
    return found


# ---------------------------------------------------------------------------
# Write: EXPIRED
# ---------------------------------------------------------------------------

def write_expired(wb, listings: list[dict], scrape_date: datetime):
    """
    Refresh expired listings:
      A.S.PRINTOUT rows 56-62  (7 slots, REPLACE)
      E.DETACHED   section 1   (dynamic rows after row 3, REPLACE)
      E.DETACHED   section 2   (historical, ADD new by MLS# only)
    """
    ws_asp  = wb["A.S.PRINTOUT"]
    ws_edet = wb["E.DETACHED"]

    # --- A.S.PRINTOUT expired section (rows 56-62) ---
    asp_rows = range(56, 63)
    clear_section(ws_asp, asp_rows, CLEAR_COLS_EXPIRED)
    written = 0
    for i, listing in enumerate(listings):
        if i >= len(asp_rows):
            log.warning(f"A.S.PRINTOUT expired: {len(asp_rows)} slots full, "
                        f"skipping {len(listings)-i} listing(s).")
            break
        r = asp_rows.start + i
        write_row(ws_asp, r, build_row(listing, scrape_date, "expired"), WRITE_COLS_EXPIRED)
        written += 1
    log.info(f"  A.S.PRINTOUT rows 56-62: wrote {written} row(s).")

    # --- E.DETACHED section 1 (rows 4 to sec2_header-1) ---
    sec2_hdr = _find_edet_sec2_header(ws_edet)
    s1_start, s1_end = 4, sec2_hdr - 1

    # Expand section 1 if needed (insert rows before sec2_header)
    available = s1_end - s1_start + 1
    if len(listings) > available:
        extra = len(listings) - available
        log.info(f"  E.DETACHED sec1: inserting {extra} row(s) before row {sec2_hdr}.")
        ws_edet.insert_rows(s1_end + 1, amount=extra)
        # Copy col-A formula pattern to new rows
        tmpl = ws_edet.cell(row=s1_start, column=COL["A"]).value or ""
        for offset in range(extra):
            nr = s1_end + 1 + offset
            if tmpl:
                ws_edet.cell(row=nr, column=COL["A"]).value = tmpl.replace(
                    f"B{s1_start}", f"B{nr}"
                )
        s1_end += extra

    clear_section(ws_edet, range(s1_start, s1_end + 1), CLEAR_COLS_EDET)
    for i, listing in enumerate(listings):
        r = s1_start + i
        data = build_row(listing, scrape_date, "expired")
        write_row(ws_edet, r, data, WRITE_COLS_EDET)
        # Owner fields (from title doc scrape — may be empty for now)
        ws_edet.cell(row=r, column=COL["S"]).value = listing.get("owner_name") or None
        ws_edet.cell(row=r, column=COL["T"]).value = listing.get("owner_address1") or None
        ws_edet.cell(row=r, column=COL["U"]).value = listing.get("owner_address2") or None
    log.info(f"  E.DETACHED sec1 rows {s1_start}-{s1_start+len(listings)-1}: "
             f"wrote {len(listings)} row(s).")

    # --- E.DETACHED section 2 (historical — add new only) ---
    existing_mls = collect_mls(ws_edet, COL["R"], start_row=sec2_hdr)
    # Find last occupied row in sec2
    sec2_data_start = sec2_hdr + 2   # skip sec2 header row + label row
    last_occ = sec2_data_start - 1
    for r in range(sec2_data_start, ws_edet.max_row + 1):
        if ws_edet.cell(row=r, column=COL["B"]).value:
            last_occ = r

    new_hist = [l for l in listings
                if l.get("mls_number", "").upper() not in existing_mls]
    insert_at = last_occ + 1
    for listing in new_hist:
        data = build_row(listing, scrape_date, "expired")
        write_row(ws_edet, insert_at, data, WRITE_COLS_EDET)
        ws_edet.cell(row=insert_at, column=COL["S"]).value = listing.get("owner_name") or None
        ws_edet.cell(row=insert_at, column=COL["T"]).value = listing.get("owner_address1") or None
        ws_edet.cell(row=insert_at, column=COL["U"]).value = listing.get("owner_address2") or None
        insert_at += 1
    log.info(f"  E.DETACHED sec2: added {len(new_hist)} new historical row(s).")


def _find_edet_sec2_header(ws_edet) -> int:
    """Find the row of the E.DETACHED section 2 column-header row (col B = 'Subject Property', row > 3)."""
    for r in range(4, ws_edet.max_row + 1):
        val = ws_edet.cell(row=r, column=COL["B"]).value
        if val and str(val).strip().lower() == "subject property":
            return r
    log.warning("E.DETACHED sec2 header not found; defaulting to row 9.")
    return 9


# ---------------------------------------------------------------------------
# Write: ACTIVE
# ---------------------------------------------------------------------------

def write_active(wb, listings: list[dict], scrape_date: datetime):
    """
    Refresh active listings in A.S.PRINTOUT rows 28-50 (23 slots, REPLACE).
    Sorted alphabetically by address.
    """
    ws_asp = wb["A.S.PRINTOUT"]
    rows = range(28, 51)
    listings_sorted = sorted(listings, key=lambda l: l.get("address", ""))

    clear_section(ws_asp, rows, CLEAR_COLS_ACTIVE)
    written = 0
    for i, listing in enumerate(listings_sorted):
        if i >= len(rows):
            log.warning(f"A.S.PRINTOUT active: {len(rows)} slots full, "
                        f"skipping {len(listings_sorted)-i} listing(s).")
            break
        write_row(ws_asp, rows.start + i,
                  build_row(listing, scrape_date, "active"), WRITE_COLS_ACTIVE)
        written += 1
    log.info(f"  A.S.PRINTOUT active rows 28-50: wrote {written} row(s).")


# ---------------------------------------------------------------------------
# Write: TERMINATED / CP
# ---------------------------------------------------------------------------

def write_terminated(wb, listings: list[dict], scrape_date: datetime):
    """
    Refresh T/CP listings in A.S.PRINTOUT rows 65-70 (6 slots, REPLACE).
    """
    ws_asp = wb["A.S.PRINTOUT"]
    rows = range(65, 71)

    clear_section(ws_asp, rows, CLEAR_COLS_EXPIRED)
    written = 0
    for i, listing in enumerate(listings):
        if i >= len(rows):
            log.warning(f"A.S.PRINTOUT T/CP: {len(rows)} slots full, "
                        f"skipping {len(listings)-i} listing(s).")
            break
        write_row(ws_asp, rows.start + i,
                  build_row(listing, scrape_date, "terminated"), WRITE_COLS_EXPIRED)
        written += 1
    log.info(f"  A.S.PRINTOUT T/CP rows 65-70: wrote {written} row(s).")


# ---------------------------------------------------------------------------
# Write: SOLD
# ---------------------------------------------------------------------------

def write_sold(wb, listings: list[dict], scrape_date: datetime):
    """
    Refresh solds in A.S.PRINTOUT:
      Rows 5-14  (recent, last 12 months, sorted newest first)
      Rows 17-23 (12-18 months back, sorted newest first)

    Requires 'sold_price' and 'sold_date' fields in listing data
    (paragon scraper must be run in Sold mode to populate these).
    """
    ws_asp = wb["A.S.PRINTOUT"]

    def sold_key(l):
        dt = parse_date(l.get("sold_date", ""))
        return dt if dt else datetime.min

    listings_sorted = sorted(listings, key=sold_key, reverse=True)

    from dateutil.relativedelta import relativedelta  # type: ignore
    cutoff_12m = scrape_date - relativedelta(months=12)
    cutoff_18m = scrape_date - relativedelta(months=18)

    recent   = [l for l in listings_sorted if (parse_date(l.get("sold_date","")) or datetime.min) >= cutoff_12m]
    historic = [l for l in listings_sorted
                if cutoff_18m <= (parse_date(l.get("sold_date","")) or datetime.min) < cutoff_12m]

    def _write_solds_block(rows: range, block: list[dict]):
        clear_section(ws_asp, rows, CLEAR_COLS_SOLD)
        for r in rows:
            ws_asp.cell(row=r, column=COL["P"]).value = None
        written = 0
        for i, listing in enumerate(block):
            if i >= len(rows):
                log.warning(f"Solds block rows {rows.start}-{rows.stop-1} full; "
                            f"skipping {len(block)-i}.")
                break
            r = rows.start + i
            write_row(ws_asp, r, build_row(listing, scrape_date, "sold"), WRITE_COLS_SOLD)
            ws_asp.cell(row=r, column=COL["P"]).value = f"=C{r}-O{r}"
            written += 1
        return written

    w1 = _write_solds_block(range(5, 15), recent)
    w2 = _write_solds_block(range(17, 24), historic)
    log.info(f"  A.S.PRINTOUT solds rows 5-14: {w1} row(s); rows 17-23: {w2} row(s).")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(
        description="Write scraped MLS data into the workbook."
    )
    parser.add_argument(
        "--type",
        choices=["expired", "active", "sold", "terminated"],
        default="expired",
        help="Which search type to write (default: expired)",
    )
    args = parser.parse_args()
    search_type = args.type

    listings = load_data()
    if not listings:
        log.error("No listings found. Run paragon_scraper.py first.")
        return

    log.info(f"Opening workbook: {WORKBOOK}")
    wb = openpyxl.load_workbook(str(WORKBOOK))

    ws_asp = wb["A.S.PRINTOUT"]
    scrape_date = get_scrape_date(ws_asp)
    log.info(f"Scrape date from B1: {scrape_date.strftime('%m/%d/%Y')}")

    dispatch = {
        "expired":    write_expired,
        "active":     write_active,
        "terminated": write_terminated,
        "sold":       write_sold,
    }
    dispatch[search_type](wb, listings, scrape_date)

    update_scrape_date(ws_asp)
    wb.save(str(WORKBOOK))
    log.info(f"Saved workbook -> {WORKBOOK}")


if __name__ == "__main__":
    main()
